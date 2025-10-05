# -*- coding: utf-8 -*-
"""
Modified on Thu Sep 26 2025

@author: yuxin
修改说明：
1. 输入：从Excel中读取关键词列表（每行一个关键词/动漫名）。
2. 输出：单个Excel，字段仅保留 Num, title, Price, Deal, Keyword。
"""

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pyquery import PyQuery as pq
import time
import openpyxl as op
import pandas as pd

# --- 全局变量 ---
driver = None
wait = None
AUTOMATION_MODE = "2"  # 默认为稳健模式

def init_browser(chromedriver_path):
    """初始化浏览器"""
    global driver, wait
    try:
        service = Service(executable_path=chromedriver_path)
        options = webdriver.ChromeOptions()
        options.add_experimental_option("excludeSwitches", ['enable-automation'])
        driver = webdriver.Chrome(service=service, options=options)
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",
                               {"source": """Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"""})
        driver.maximize_window()
        wait = WebDriverWait(driver, 20)
        print("浏览器启动成功！")
        return True
    except Exception as e:
        print(f"浏览器启动失败，请检查路径: {e}")
        return False

def user_confirm(message):
    if AUTOMATION_MODE == "1":
        input(message)
    else:
        print(message + " (自动跳过)")
        time.sleep(2)

def search_goods(keyword):
    """搜索关键词"""
    try:
        driver.get('https://s.taobao.com/')
        print(f"正在搜索: {keyword}")
        input_box = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="q"]')))
        submit_btn = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="J_SearchForm"]/div/div[1]/button')))
        input_box.clear()
        input_box.send_keys(keyword)
        submit_btn.click()
        time.sleep(2)
        return True
    except Exception as exc:
        print(f"search_goods函数错误！Error：{exc}")
        return False

def page_turning(page_number):
    """翻页"""
    try:
        print(f"正在翻页: 第 {page_number} 页")
        # 查找“下一页”按钮，注意淘宝的控件可能是 button 或 a 标签
        next_page_buttons = driver.find_elements(By.XPATH, "//button[span[text()='下一页']] | //a[span[text()='下一页']]")
        if not next_page_buttons:
            print(f"没有找到第 {page_number} 页的下一页按钮，停止翻页。")
            return False  # 翻页失败
        next_page_button = next_page_buttons[0]
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_page_button)
        next_page_button.click()
        # 页面加载时间
        time.sleep(2)
        return True  # 翻页成功
    except Exception as exc:
        print(f"翻页出错: {exc}")
        return False




def get_goods(page, worksheet, count, keyword):
    """获取商品信息，只保留核心字段"""
    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.content--CUnfXXxv")))
        if AUTOMATION_MODE == "2":
            time.sleep(1.5)

        html = driver.page_source
        doc = pq(html)
        items = list(doc('div.content--CUnfXXxv > div > div').items())

        for item in items:
            if not item.find('.title--qJ7Xg_90 span').text():
                continue

            title = item.find('.title--qJ7Xg_90 span').text()
            price_str = item.find('.priceInt--yqqZMJ5a').text() + item.find('.priceFloat--XpixvyQ1').text()
            price = float(price_str) if price_str else 0.0
            deal_str = item.find('.realSales--XZJiepmt').text()
            deal = deal_str.replace("万", "0000").replace("+", "").replace("人付款", "").strip()
            deal = int(float(deal)) if deal else 0

            # 写入Excel（只保留 Num, title, Price, Deal, Keyword）
            worksheet.cell(row=count, column=1, value=count - 1)
            worksheet.cell(row=count, column=2, value=title)
            worksheet.cell(row=count, column=3, value=price)
            worksheet.cell(row=count, column=4, value=deal)
            worksheet.cell(row=count, column=5, value=keyword)

            count += 1
        return count
    except Exception as exc:
        print(f"get_goods函数错误！Error：{exc}")
        return count

def crawl_keyword(keyword, start_page, end_page, worksheet, start_count):
    """爬取单个关键词"""
    if not search_goods(keyword):
        print(f"搜索关键词 '{keyword}' 失败，跳过。")
        return start_count

    excel_count = start_count
    try:
        print(f"\n--- 开始爬取: '{keyword}', 第 {start_page} 页 ---")
        user_confirm('确认界面加载完毕, 按下回车...')
        excel_count = get_goods(start_page, worksheet, excel_count, keyword)
        for i in range(start_page + 1, end_page + 1):
            if not page_turning(i):  # 没有下一页就停止
                break
            print(f"\n--- 开始爬取: '{keyword}', 第 {i} 页 ---")
            # 全自动模式下无需人工确认
            user_confirm('确认界面加载完毕, 按下回车...')
            excel_count = get_goods(i, worksheet, excel_count, keyword)




    except Exception as exc:
        print(f"爬取 '{keyword}' 出错！Error：{exc}")
    return excel_count

if __name__ == '__main__':
    CHROME_DRIVER_PATH = r"C:/Users/HP/Code/chromedriver-win64/chromedriver-win64/chromedriver.exe"  # 修改为你的 chromedriver 路径
    INPUT_EXCEL = "newall_cleaned.xlsx"  # 你的输入文件（第一列是关键词）
    pageStart = 1
    pageEnd = 15

    # 模式选择
    while True:
        mode = input("请选择模式：1.稳健模式  2.全自动模式：")
        if mode in ["1", "2"]:
            AUTOMATION_MODE = mode
            break

    # 读取Excel关键词
    df = pd.read_excel(INPUT_EXCEL, sheet_name="anime")

    # 获取 anime 列的非空关键词列表
    keywords_list = df['anime'].dropna().tolist()

    # 创建结果Excel
    workbook = op.Workbook()
    worksheet = workbook.active
    worksheet.title = "TaobaoData"
    worksheet.append(["Num", "title", "Price", "Deal", "Keyword"])
    count = 2

    if init_browser(CHROME_DRIVER_PATH):
        for keyword in keywords_list:
            print(f"========== 开始处理: [{keyword}] ==========")
            count = crawl_keyword(keyword, pageStart, pageEnd, worksheet, count)

        if driver:
            driver.quit()

    # 保存汇总结果
    data = time.strftime('%Y%m%d-%H%M', time.localtime(time.time()))
    filename = f"Taobao_All_{data}.xlsx"
    workbook.save(filename=filename)
    print(f"所有关键词数据已保存至: {filename}")
