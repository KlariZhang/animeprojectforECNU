# -*- coding: utf-8 -*-
"""
Created on Fri Aug 29 17:39:08 2025

@author: yuxin
"""

# 代码说明：
'''
代码功能： 基于ChromeDriver爬取taobao（淘宝）平台商品列表数据 (批量关键词 + 双模式版)
模式选择： 1. 稳健模式 (每次翻页和新任务前需手动按回车确认，适合处理验证码)
          2. 全自动模式 (全程无需干预，自动执行所有任务)
'''
# 声明第三方库/头文件
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pyquery import PyQuery as pq
import time
import openpyxl as op

# --- 全局变量 ---
driver = None
wait = None
AUTOMATION_MODE = "1" # 默认为稳健模式

def init_browser(chromedriver_path):
    """初始化并返回一个配置好的浏览器实例"""
    global driver, wait
    try:
        service = Service(executable_path=chromedriver_path)
        options = webdriver.ChromeOptions()
        options.add_experimental_option("excludeSwitches", ['enable-automation'])
        driver = webdriver.Chrome(service=service, options=options)
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",
                               {"source": """Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"""})
        driver.maximize_window()
        wait = WebDriverWait(driver, 20)
        print("浏览器启动成功！")
        return True
    except Exception as e:
        print(f"浏览器启动失败，请检查ChromeDriver路径是否正确: {e}")
        return False

def user_confirm(message):
    """根据模式决定是否需要用户手动确认"""
    if AUTOMATION_MODE == "1":
        input(message)
    else:
        print(message + " (自动跳过)")
        time.sleep(2) # 自动模式下用短暂延时代替手动等待

def search_goods(keyword):
    """输入“关键词”，搜索"""
    try:
        driver.get('https://s.taobao.com/')
        print(f"正在搜索: {keyword}")
        input_box = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="q"]')))
        submit_btn = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="J_SearchForm"]/div/div[1]/button')))
        input_box.clear()
        input_box.send_keys(keyword)
        submit_btn.click()
        time.sleep(2)
        print("搜索完成！")
        return True
    except Exception as exc:
        print(f"search_goods函数错误！Error：{exc}")
        return False

def turn_pageStart(start_page):
    """翻页至第start_page页，使用更健壮的定位器"""
    if start_page == 1:
        return
    try:
        print(f"正在翻转至第 {start_page} 页")
        page_input_xpath = "//div[contains(@class, 'J_AjaxForm')]//input"
        page_input_element = wait.until(EC.presence_of_element_located((By.XPATH, page_input_xpath)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", page_input_element)
        time.sleep(1)
        page_input_element.clear()
        page_input_element.send_keys(start_page)
        admit_button_xpath = "//button[span[text()='确定']]"
        admit = wait.until(EC.element_to_be_clickable((By.XPATH, admit_button_xpath)))
        admit.click()
        print(f"已翻至第 {start_page} 页")
    except Exception as exc:
        print(f"turn_pageStart函数错误！Error：{exc}")

def page_turning(page_number):
    """翻页函数，使用更健壮的定位器"""
    try:
        print(f"正在翻页: 第 {page_number} 页")
        next_page_xpath = "//button[span[text()='下一页']]"
        next_page_button = wait.until(EC.presence_of_element_located((By.XPATH, next_page_xpath)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_page_button)
        time.sleep(1)
        wait.until(EC.element_to_be_clickable(next_page_button)).click()
        current_page_xpath = "//span[contains(@class, 'currentPage')]/em"
        wait.until(EC.text_to_be_present_in_element((By.XPATH, current_page_xpath), str(page_number)))
        print(f"已翻至: 第 {page_number} 页")
    except Exception as exc:
        print(f"page_turning函数错误！Error：{exc}")

def get_goods(page, worksheet, count):
    """获取每一页的商品信息"""
    try:
        # 智能等待商品列表出现
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.content--CUnfXXxv")))
        
        # 在全自动模式下，多给一点渲染时间
        if AUTOMATION_MODE == "2":
            time.sleep(1.5)
            
        html = driver.page_source
        doc = pq(html)
        items = list(doc('div.content--CUnfXXxv > div > div').items())
        
        if not items:
            print("警告：当前页面未找到商品列表，可能出现了验证码或页面结构已更改。")
            user_confirm("请检查页面是否正常，手动处理后按“Enter”重试...")
            # 重试一次
            html = driver.page_source
            doc = pq(html)
            items = list(doc('div.content--CUnfXXxv > div > div').items())

        for item in items:
            # 过滤无效的div块
            if item.find('.title--RoseSo8H').text() == '大家都在搜' or \
               item.find('.headTitleText--hxVemljn').text() == '对本次搜索体验满意吗' or \
               not item.find('.title--qJ7Xg_90 span').text():
                continue

            title = item.find('.title--qJ7Xg_90 span').text()
            price_str = item.find('.priceInt--yqqZMJ5a').text() + item.find('.priceFloat--XpixvyQ1').text()
            price = float(price_str) if price_str else 0.0
            deal_str = item.find('.realSales--XZJiepmt').text()
            deal = deal_str.replace("万", "0000").replace("+", "").replace("人付款", "").strip()
            deal = int(float(deal)) if deal else 0
            location = item.find('.procity--wlcT2xH9 span').text()
            shop = item.find('.shopNameText--DmtlsDKm').text()
            postText = item.find('.subIconWrapper--Vl8zAdQn').text()
            isPostFree = "包邮" if "包邮" in postText else "不包邮"
            t_url = item.find('.doubleCardWrapperAdapt--mEcC7olq').attr('href')
            shop_url = item.find('.TextAndPic--grkZAtsC a').attr('href')
            img_url = item.find('.mainPicAdaptWrapper--V_ayd2hD img').attr('src')
            
            product = {
                'Page': page, 'Num': count - 1, 'title': title, 'price': price,
                'deal': deal, 'location': location, 'shop': shop,
                'isPostFree': isPostFree, 'url': t_url, 'shop_url': shop_url, 'img_url': img_url
            }
            print(product)
            
            # 写入Excel
            worksheet.cell(row=count, column=1, value=count - 1)
            worksheet.cell(row=count, column=2, value=title)
            worksheet.cell(row=count, column=3, value=price)
            worksheet.cell(row=count, column=4, value=deal)
            worksheet.cell(row=count, column=5, value=location)
            worksheet.cell(row=count, column=6, value=shop)
            worksheet.cell(row=count, column=7, value=isPostFree)
            worksheet.cell(row=count, column=8, value=t_url)
            worksheet.cell(row=count, column=9, value=shop_url)
            worksheet.cell(row=count, column=10, value=img_url)
            count += 1
        
        return count
    except Exception as exc:
        print(f"get_goods函数错误！Error：{exc}")
        return count

def crawl_keyword(keyword, start_page, end_page):
    """封装了爬取单个关键词的完整流程"""
    try:
        workbook = op.Workbook()
        worksheet = workbook.active
        worksheet.title = keyword[:30]
        title_list = ['Num', 'title', 'Price', 'Deal', 'Location', 'Shop', 'IsPostFree', 'Title_URL', 'Shop_URL', 'Img_URL']
        worksheet.append(title_list)
        excel_count = 2
    except Exception as exc:
        print(f"为关键词 '{keyword}' 创建Excel失败！Error：{exc}")
        return

    if not search_goods(keyword):
        print(f"搜索关键词 '{keyword}' 失败，跳过此任务。")
        return

    turn_pageStart(start_page)
    
    try:
        # 爬取起始页
        print(f"\n--- 开始爬取关键词: '{keyword}', 第 {start_page} 页 ---")
        user_confirm('确认界面加载完毕(处理登录或验证码), 按下“Enter”开始爬取...')
        excel_count = get_goods(start_page, worksheet, excel_count)
        
        # 爬取后续页面
        for i in range(start_page + 1, end_page + 1):
            page_turning(i)
            print(f"\n--- 开始爬取关键词: '{keyword}', 第 {i} 页 ---")
            user_confirm('确认界面加载完毕, 按下“Enter”开始爬取...')
            excel_count = get_goods(i, worksheet, excel_count)
    except Exception as exc:
        print(f"爬取关键词 '{keyword}' 过程中发生错误！Error：{exc}")
    finally:
        data = time.strftime('%Y%m%d-%H%M', time.localtime(time.time()))
        safe_keyword = "".join(x for x in keyword if x.isalnum() or x in " _-")
        filename = f"{safe_keyword}_{data}_FromTB.xlsx"
        workbook.save(filename=filename)
        print(f"\n关键词 '{keyword}' 的数据已成功保存至: {filename}\n")

if __name__ == '__main__':
    # --- 用户输入配置 ---
    CHROME_DRIVER_PATH = r"C:\Program Files\Google\Chrome\Application\chromedriver-win64\chromedriver-win64 (1)\chromedriver-win64\chromedriver.exe"
    KEYWORDS_INPUT = input('输入搜索的商品关键词(用英文逗号,隔开)：')
    pageStart = int(input('输入爬取的起始页PageStart：'))
    pageEnd = int(input('输入爬取的终止页PageEnd：'))
    
    # [ADDED] 模式选择
    while True:
        mode = input("请选择运行模式：\n 1. 稳健模式 (需手动确认, 适合处理验证码)\n 2. 全自动模式 (全程无需干预)\n请输入模式编号 (1 或 2): ")
        if mode in ["1", "2"]:
            AUTOMATION_MODE = mode
            break
        else:
            print("输入无效，请输入 1 或 2。")

    keywords_list = [kw.strip() for kw in KEYWORDS_INPUT.split(',') if kw.strip()]
    
    if not keywords_list:
        print("未输入有效的关键词，程序退出。")
    else:
        if init_browser(CHROME_DRIVER_PATH):
            for keyword in keywords_list:
                print(f"========== 开始处理关键词: [{keyword}] ==========")
                crawl_keyword(keyword, pageStart, pageEnd)
            
            if driver:
                driver.quit()
            print("========== 所有任务已完成 ==========")